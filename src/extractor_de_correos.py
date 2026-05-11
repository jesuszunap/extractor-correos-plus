import os
import re
import json
import time
import logging
import threading
import sys
import unicodedata
import win32com.client
import pywintypes
import pythoncom
import pandas as pd

from datetime import datetime, timedelta
from pathlib import Path
from calendar import monthrange

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from colorama import init, Fore


def cargar_pywinauto_desktop():
    tenia_frozen = hasattr(sys, "frozen")
    frozen_original = getattr(sys, "frozen", None)

    try:
        sys.frozen = "console_exe"
        from pywinauto import Desktop, mouse
        return Desktop, mouse
    except ImportError:
        return None, None
    except Exception:
        logging.getLogger("extractor_correos").exception("No se pudo cargar pywinauto")
        return None, None
    finally:
        if tenia_frozen:
            sys.frozen = frozen_original
        else:
            try:
                delattr(sys, "frozen")
            except AttributeError:
                pass


Desktop, pywinauto_mouse = cargar_pywinauto_desktop()


# ============================================================
# ConfiguraciÃ³n general
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
PERSONAS_JSON = BASE_DIR / "personas.json"
LOG_DIR = PROJECT_DIR / "logs"
LOG_PATH = LOG_DIR / "extractor_correos.log"

LOG_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    encoding="utf-8"
)
logger = logging.getLogger("extractor_correos")

REMITENTES_OMITIDOS = [
    "QUIPUX",
    "UNIVERSIDAD DE GUAYAQUIL",
    "INFO UG",
    "Comunicados DVSBE",
    "Zoom",
    "Titulares EL UNIVERSO",
    "Canva",
    "ClickUp Notifications",
    "ClickUp Team",
    "DepositPhotos",
]

MESES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

COM_HRESULTS_REINTENTABLES = {
    -2147418111,  # Call was rejected by callee.
    -2147417846,  # Application is busy.
    -2147023174,  # RPC server unavailable.
    -2146959355,  # Server execution failed.
    -2147352567,  # Outlook exception, often "not connected".
    -2079063787,  # OLE/MAPI transient error seen with Outlook/Word.
}


# ============================================================
# Resultado de exportaciÃ³n
# ============================================================

class ResultadoExportacion:
    def __init__(
        self,
        exito=False,
        cantidad=0,
        carpeta=None,
        excel=None,
        mensaje="",
        total=0,
        errores=0,
        omitidos=0
    ):
        self.exito = exito
        self.cantidad = cantidad
        self.carpeta = carpeta
        self.excel = excel
        self.mensaje = mensaje
        self.total = total
        self.errores = errores
        self.omitidos = omitidos


# ============================================================
# Utilidades de texto
# ============================================================

def limpiar_acortar_remitentes(texto):
    if not texto:
        return ""
    texto = str(texto).strip()
    texto = re.sub(r'[\\/:*?"<>|\r\n\tâ€œâ€â€˜â€™Â´`]', "_", texto)
    texto = re.sub(r'\s+', " ", texto).strip()
    if len(texto) > 45:
        texto = texto[:45]
    return f"{texto}"


def limpiar_texto(nombre):
    if not nombre:
        return ""
    nombre = str(nombre).strip()

    m = re.match(r"(.+)\.([A-Za-z0-9]{1,5})$", nombre)
    if m:
        base, ext = m.group(1), "." + m.group(2).lower()
    else:
        base, ext = nombre, ""

    base = re.sub(r'[\\/:*?"<>|\r\n\tâ€œâ€â€˜â€™Â´`]', "_", base)
    base = re.sub(r'\s+', " ", base).strip()

    if len(base) > 70:
        base = base[:70]

    return f"{base}{ext.lower()}"


def quitar_acentos(texto):
    if not texto:
        return ""
    texto = str(texto)
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )


def limpiar_nombre(nombre):
    if not nombre:
        return ""

    nombre = quitar_acentos(str(nombre).lower())
    nombre = re.sub(
        r'\b(arq|arquitecto|arquitecta|ing|ingeniero|ingeniera|lic|licenciado|licenciada|sr|sra|srta|senor|senora|senorita|dra|dr|ab|abogado|abogada|mgs|magister|phd|ph\.d)\.?',
        '',
        nombre
    )
    nombre = nombre.replace('.', '').replace(',', '')
    nombre = re.sub(r'\s+', ' ', nombre).strip()
    return nombre


def nompropio_python(texto):
    if texto is None:
        return ""
    return str(texto).title()


def normalizar_email(email):
    if not email:
        return ""
    return str(email).strip().replace(" ", "").replace(",", ".").lower()


def es_error_com_reintentable(error):
    if not isinstance(error, pywintypes.com_error):
        return False

    hresult = error.args[0] if error.args else None
    return hresult in COM_HRESULTS_REINTENTABLES


def mensaje_error_com(error):
    hresult = error.args[0] if getattr(error, "args", None) else None

    if hresult in {-2146959355, -2079063787, -2147352567}:
        return (
            "Outlook o Word no respondio a tiempo. Si Outlook muestra una ventana "
            "preguntando si desea conectar, pulse Conectar y vuelva a intentar."
        )

    if hresult in {-2147418111, -2147417846}:
        return (
            "Outlook o Word esta ocupado. Cierre cuadros de dialogo abiertos y vuelva a intentar."
        )

    return str(error)


def ejecutar_com_con_reintentos(
    accion,
    descripcion,
    intentos=4,
    espera=1.5,
    progress_callback=None
):
    ultimo_error = None

    for intento in range(1, intentos + 1):
        try:
            return accion()
        except pywintypes.com_error as error:
            ultimo_error = error
            if not es_error_com_reintentable(error) or intento == intentos:
                raise

            notificar_progreso(
                progress_callback,
                f"{descripcion}: Outlook/Word esta ocupado. Reintentando {intento + 1} de {intentos}..."
            )

            logger.warning(
                "%s fallo por COM ocupado. Reintento %s de %s. Error: %s",
                descripcion,
                intento + 1,
                intentos,
                error
            )
            time.sleep(espera * intento)

    raise ultimo_error


def notificar_progreso(progress_callback, mensaje, porcentaje=None):
    if not progress_callback:
        return

    progress_callback({
        "mensaje": mensaje,
        "porcentaje": porcentaje
    })


def intentar_pulsar_conectar_outlook():
    if Desktop is None or pywinauto_mouse is None:
        logger.warning("pywinauto no esta instalado. No se puede auto-pulsar Conectar en Outlook.")
        return False

    for backend in ("uia", "win32"):
        try:
            escritorio = Desktop(backend=backend)
            ventanas = escritorio.windows(title_re=".*Outlook.*", visible_only=True)

            for ventana in ventanas:
                try:
                    textos = ventana.texts()
                    texto_ventana = quitar_acentos(" ".join(textos)).lower()
                    titulo = ventana.window_text()

                    if "microsoft outlook" not in quitar_acentos(titulo).lower():
                        continue

                    es_dialogo_conexion = (
                        "conexion de uso medido" in texto_ventana
                        or ("conectar" in texto_ventana and "salir de outlook" in texto_ventana)
                    )

                    es_dialogo_outlook_ciego = (
                        titulo == "Microsoft Outlook"
                        and len(textos) == 1
                        and textos[0] == "Microsoft Outlook"
                    )

                    if es_dialogo_outlook_ciego:
                        rect = ventana.rectangle()
                        x = rect.left + int(rect.width() * 0.29)
                        y = rect.bottom - 20
                        ventana.set_focus()
                        time.sleep(0.2)
                        pywinauto_mouse.click(button="left", coords=(x, y))
                        logger.info(
                            "pywinauto/%s pulso Conectar por coordenadas en dialogo Outlook ciego. Rect=%s",
                            backend,
                            rect
                        )
                        return True

                    if not es_dialogo_conexion:
                        logger.info(
                            "pywinauto/%s vio ventana Outlook no compatible. Titulo=%s Textos=%s",
                            backend,
                            titulo,
                            textos
                        )
                        continue

                    logger.info(
                        "pywinauto/%s detecto advertencia de conexion Outlook. Titulo=%s Textos=%s",
                        backend,
                        titulo,
                        textos
                    )

                    if backend == "uia":
                        boton = ventana.child_window(title_re=".*Conectar.*", control_type="Button")
                    else:
                        boton = ventana.child_window(title_re=".*Conectar.*", class_name="Button")

                    if not boton.exists(timeout=0.5):
                        logger.info("pywinauto/%s no encontro boton Conectar.", backend)
                        continue

                    ventana.set_focus()
                    time.sleep(0.2)
                    boton.click_input()
                    logger.info("pywinauto/%s pulso Conectar en advertencia de Outlook.", backend)
                    return True

                except Exception:
                    logger.exception("Error intentando pulsar Conectar con pywinauto/%s", backend)

        except Exception:
            logger.exception("Error buscando ventana de conexion Outlook con pywinauto/%s", backend)

    return False


def iniciar_auto_conectar_outlook(progress_callback=None, duracion_segundos=45):
    def vigilar():
        pythoncom.CoInitialize()
        logger.info("Vigilante de conexion Outlook iniciado.")

        try:
            fin = time.time() + duracion_segundos

            while time.time() < fin:
                if intentar_pulsar_conectar_outlook():
                    notificar_progreso(
                        progress_callback,
                        "Outlook pidio conectar por conexion medida. Se pulso Conectar automaticamente."
                    )
                    return

                time.sleep(0.5)
        finally:
            pythoncom.CoUninitialize()
            logger.info("Vigilante de conexion Outlook finalizado.")

    hilo = threading.Thread(target=vigilar, daemon=True)
    hilo.start()


def texto_normalizado(texto):
    texto = quitar_acentos(str(texto or "").lower())
    texto = re.sub(r'[^a-z0-9@._\s]', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto).strip()
    return texto


# ============================================================
# Carga y bÃºsqueda de personas desde personas.json
# ============================================================

PERSONAS = []


def cargar_personas():
    if not PERSONAS_JSON.exists():
        return []

    try:
        with open(PERSONAS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, dict):
            personas = data.get("personas", [])
        elif isinstance(data, list):
            personas = data
        else:
            personas = []

        for persona in personas:
            persona["_nombre_limpio"] = limpiar_nombre(persona.get("nombre_completo", ""))
            persona["_nombre_corto_limpio"] = limpiar_nombre(persona.get("nombre_corto", ""))
            persona["_email_limpio"] = normalizar_email(persona.get("email", ""))

        return personas

    except Exception:
        return []


def inicializar_personas():
    global PERSONAS
    PERSONAS = cargar_personas()
    return PERSONAS


def fecha_en_vigencia(persona, fecha_correo):
    """
    Verifica si la fecha del correo cae dentro de la vigencia del cargo.
    Si no hay fechas, el registro queda disponible como fallback.
    """
    desde = persona.get("vigente_desde")
    hasta = persona.get("vigente_hasta")

    try:
        if desde:
            desde_dt = datetime.strptime(desde, "%Y-%m-%d")
            if fecha_correo < desde_dt:
                return False

        if hasta:
            hasta_dt = datetime.strptime(hasta, "%Y-%m-%d")
            if fecha_correo > hasta_dt:
                return False

        return True

    except Exception:
        return True


def puntuar_persona(persona, texto):
    texto_original = str(texto or "")
    texto_limpio = limpiar_nombre(texto_original)
    texto_email = normalizar_email(texto_original)

    if not texto_limpio and not texto_email:
        return 0

    score = 0

    email = persona.get("_email_limpio", "")
    if email and email in texto_email:
        score += 100

    nombre_limpio = persona.get("_nombre_limpio", "")
    nombre_corto_limpio = persona.get("_nombre_corto_limpio", "")

    if nombre_limpio and nombre_limpio in texto_limpio:
        score += 80

    if nombre_corto_limpio and nombre_corto_limpio in texto_limpio:
        score += 60

    palabras_persona = set((nombre_limpio + " " + nombre_corto_limpio).split())
    palabras_texto = set(texto_limpio.split())
    coincidencias = palabras_persona.intersection(palabras_texto)

    score += len(coincidencias) * 12

    if persona.get("activo") is True:
        score += 5
    elif persona.get("activo") is False:
        score += 1

    return score


def buscar_persona(texto, fecha_correo=None, minimo=24):
    mejor = None
    mejor_score = 0

    for persona in PERSONAS:
        if fecha_correo and not fecha_en_vigencia(persona, fecha_correo):
            continue

        score = puntuar_persona(persona, texto)
        if score > mejor_score:
            mejor = persona
            mejor_score = score

    if mejor and mejor_score >= minimo:
        return mejor

    return None


def formatear_persona(persona, usar_corto=False):
    if not persona:
        return ""

    titulo = persona.get("titulo_abreviado") or ""
    nombre = persona.get("nombre_corto") if usar_corto else persona.get("nombre_completo")
    nombre = nombre or persona.get("nombre_corto") or ""

    if titulo and not nombre.lower().startswith(titulo.lower()):
        return f"{titulo} {nombre}".strip()

    return nombre.strip()


def nombres_conocidos_cc(cc, fecha_correo=None):
    if not cc:
        return "-----"

    partes_cc = re.split(r'[;,]', str(cc))
    resultado = []

    for parte in partes_cc:
        persona = buscar_persona(parte, fecha_correo)
        if persona:
            nombre = formatear_persona(persona, usar_corto=True)
            if nombre and nombre not in resultado:
                resultado.append(nombre)

    if len(resultado) == 0:
        return "-----"

    return "\n".join(resultado)


def nombres_conocidos_rem(rem, fecha_correo=None):
    if not rem:
        return ""

    persona = buscar_persona(rem, fecha_correo)
    if persona:
        return formatear_persona(persona, usar_corto=False)

    return rem


def cut_nombres_destinatarios(destinatario: str):
    if not destinatario:
        return []

    personas = re.split(r'[;,]', str(destinatario))
    resultado = []

    for p in personas:
        p = p.strip()
        if not p:
            continue

        p = re.sub(r"<.*?>", "", p).strip()
        partes = p.split()

        if len(partes) == 1:
            resultado.append(partes[0])
        elif len(partes) == 2:
            resultado.append(f"{partes[0]} {partes[1]}")
        else:
            resultado.append(f"{partes[0]} {partes[-2]}")

    return resultado


def obtener_info_destinatarios(lista_nombres, fecha_correo=None):
    destinatarios = []
    cargos = []

    for nombre_abreviado in lista_nombres:
        persona = buscar_persona(nombre_abreviado, fecha_correo)

        if persona:
            destinatarios.append(formatear_persona(persona, usar_corto=True))
            cargos.append(persona.get("cargo") or "")
        else:
            destinatarios.append(nombre_abreviado)

    return "\n\n".join(destinatarios), "\n\n".join([c for c in cargos if c])


def obtener_info_remitente(remitente, fecha_correo=None):
    persona = buscar_persona(remitente, fecha_correo)

    if persona:
        return persona.get("cargo") or None, persona.get("dependencia") or None

    return None, None


# ============================================================
# Excel y anexos
# ============================================================

def exportar_excel(registros, carpeta_base, fecha_inicio_str, tipo_exportacion):
    if not registros:
        return None

    df = pd.DataFrame(registros)

    sufijo = "Recibidos" if tipo_exportacion == "recibidos" else "Enviados"
    ruta_excel = carpeta_base / f"{fecha_inicio_str}_Correos{sufijo}Exportados.xlsx"

    df.to_excel(ruta_excel, index=False)

    wb = load_workbook(ruta_excel)
    ws = wb.active
    ws.title = "Recibidos" if tipo_exportacion == "recibidos" else "Enviados"

    for row in ws.iter_rows():
        for celda in row:
            celda.font = Font(name="Arial", size=12)
            celda.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")

    if ws.max_column >= 10:
        for celda in ws["J"]:
            celda.font = Font(name="Arial", size=12, color="FF0000", bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center")

    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border


    for col in ws.columns:
        column = get_column_letter(col[0].column)
        ws.column_dimensions[column].width = 20.71

    if ws.max_column >= 7:
        ws.column_dimensions["G"].width = 30.71
    if ws.max_column >= 10:
        ws.column_dimensions["J"].width = 36

    for row in ws.iter_rows():
        celda = row[0]
        ws.row_dimensions[celda.row].height = 150.04

    ultima_fila = ws.max_row
    ultima_col = ws.max_column
    rango_tabla = f"A1:{get_column_letter(ultima_col)}{ultima_fila}"

    body_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for fila in ws[rango_tabla]:
        for celda in fila:
            if celda.value is None or celda.value == "":
                celda.fill = body_fill

    tabla = Table(displayName="Exportados", ref=rango_tabla)
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    ultima_columna = get_column_letter(ws.max_column)

    for fila in range(1, ws.max_row + 1):
        celda = ws[f"{ultima_columna}{fila}"]

        celda.border = Border(
            left=celda.border.left,
            right=Side(style="thin"),
            top=celda.border.top,
            bottom=celda.border.bottom
        )

    ws.sheet_view.zoomScale = 80
    wb.save(ruta_excel)

    return ruta_excel


def obtener_anexos(anexos, carpeta_correo):
    lista_anexos = []

    cantidad_anexos = ejecutar_com_con_reintentos(
        lambda: anexos.Count,
        "Leyendo anexos"
    )

    if cantidad_anexos > 0:
        for anexo in anexos:
            nombre_archivo = limpiar_texto(anexo.FileName)
            if "image" not in nombre_archivo.lower() and "outlook" not in nombre_archivo.lower():
                carpeta_anexos = carpeta_correo / "Anexos"
                os.makedirs(carpeta_anexos, exist_ok=True)
                ruta_anexo = carpeta_anexos / nombre_archivo
                ejecutar_com_con_reintentos(
                    lambda: anexo.SaveAsFile(str(ruta_anexo)),
                    "Guardando anexo"
                )
                lista_anexos.append(nombre_archivo)

    return lista_anexos


# ============================================================
# Fechas y configuraciÃ³n de Outlook
# ============================================================

def obtener_rango_dia(dia, mes, anio):
    fecha_inicio = datetime(int(anio), int(mes), int(dia), 0, 0, 0)
    fecha_fin = fecha_inicio + timedelta(days=1)
    etiqueta = fecha_inicio.strftime("%Y-%m-%d")
    return fecha_inicio, fecha_fin, etiqueta, MESES[int(mes)], str(anio)


def obtener_rango_mes(mes, anio):
    mes = int(mes)
    anio = int(anio)
    ultimo_dia = monthrange(anio, mes)[1]
    fecha_inicio = datetime(anio, mes, 1, 0, 0, 0)
    fecha_fin = datetime(anio, mes, ultimo_dia, 0, 0, 0) + timedelta(days=1)
    etiqueta = fecha_inicio.strftime("%Y-%m")
    return fecha_inicio, fecha_fin, etiqueta, MESES[mes], str(anio)


def obtener_config_outlook(tipo_exportacion):
    if tipo_exportacion == "recibidos":
        return {
            "folder_id": 6,
            "campo_fecha": "ReceivedTime",
            "atributo_fecha": "ReceivedTime",
            "nombre_carpeta": "Correos Recibidos",
            "sort_desc": False
        }

    return {
        "folder_id": 5,
        "campo_fecha": "SentOn",
        "atributo_fecha": "SentOn",
        "nombre_carpeta": "Correos Enviados",
        "sort_desc": False
    }


def formatear_fecha_restrict_local(fecha):
    return fecha.strftime("%d/%m/%Y %H:%M")


def formatear_fecha_restrict_outlook(fecha):
    return fecha.strftime("%m/%d/%Y %I:%M %p")


def crear_filtro_fecha(campo_fecha, fecha_inicio, fecha_fin, formateador):
    return (
        f"[{campo_fecha}] >= '{formateador(fecha_inicio)}' "
        f"AND [{campo_fecha}] < '{formateador(fecha_fin)}'"
    )


def agregar_items_restringidos(destino, vistos, items, filtro, descripcion, progress_callback=None):
    cantidad_antes = len(destino)

    try:
        filtrados = ejecutar_com_con_reintentos(
            lambda: items.Restrict(filtro),
            descripcion,
            progress_callback=progress_callback
        )

        for item in filtrados:
            try:
                entry_id = ejecutar_com_con_reintentos(
                    lambda: item.EntryID,
                    "Leyendo identificador del correo",
                    progress_callback=progress_callback
                )
            except Exception:
                entry_id = f"sin-entryid-{id(item)}"

            if entry_id not in vistos:
                destino.append(item)
                vistos.add(entry_id)

        logger.info("%s aplicado. Candidatos acumulados=%s. Filtro=%s", descripcion, len(destino), filtro)
        return len(destino) - cantidad_antes

    except Exception:
        logger.exception("%s fallo. Filtro=%s", descripcion, filtro)
        return 0


def obtener_correos_candidatos(items, config, fecha_inicio, fecha_fin, progress_callback=None):
    candidatos = []
    vistos = set()

    agregados_local = agregar_items_restringidos(
        candidatos,
        vistos,
        items,
        crear_filtro_fecha(config["campo_fecha"], fecha_inicio, fecha_fin, formatear_fecha_restrict_local),
        "Filtrando correos por fecha local",
        progress_callback=progress_callback
    )

    if agregados_local > 0:
        return candidatos

    agregar_items_restringidos(
        candidatos,
        vistos,
        items,
        crear_filtro_fecha(config["campo_fecha"], fecha_inicio, fecha_fin, formatear_fecha_restrict_outlook),
        "Filtrando correos por fecha Outlook",
        progress_callback=progress_callback
    )

    return candidatos


def obtener_fecha_mail(mail, atributo_fecha):
    fecha = ejecutar_com_con_reintentos(
        lambda: getattr(mail, atributo_fecha),
        "Leyendo fecha del correo"
    )
    return datetime(fecha.year, fecha.month, fecha.day, fecha.hour, fecha.minute, fecha.second)


def obtener_nombre_para_carpeta(tipo_exportacion, remitente, destinatarios_raw):
    if tipo_exportacion == "enviados":
        destinatarios = cut_nombres_destinatarios(destinatarios_raw)
        if destinatarios:
            return destinatarios[0]
        return "Sin destinatario"

    return remitente or "Sin remitente"


def eliminar_temporal_con_reintentos(path, intentos=5, espera=0.4):
    for intento in range(1, intentos + 1):
        try:
            if path.exists():
                os.remove(path)
            return True
        except PermissionError:
            if intento == intentos:
                logger.warning("No se pudo borrar el temporal despues de %s intentos: %s", intentos, path)
                return False
            time.sleep(espera * intento)


def limpiar_carpeta_si_vacia(path):
    try:
        if path.exists() and path.is_dir() and not any(path.iterdir()):
            path.rmdir()
            return True
    except Exception:
        logger.exception("No se pudo limpiar carpeta vacia: %s", path)

    return False


# ============================================================
# ConversiÃ³n correo a PDF
# ============================================================

def convertir_correo_a_pdf(mail, word, carpeta_correo, asunto_limpio, progress_callback=None):
    mht_path = carpeta_correo / f"{asunto_limpio}.mht"
    pdf_path = carpeta_correo / f"{asunto_limpio}.pdf"
    doc = None

    try:
        ejecutar_com_con_reintentos(
            lambda: mail.SaveAs(str(mht_path), 10),
            "Guardando correo temporal",
            progress_callback=progress_callback
        )

        doc = ejecutar_com_con_reintentos(
            lambda: word.Documents.Open(
                str(mht_path),
                ConfirmConversions=False,
                ReadOnly=True,
                AddToRecentFiles=False,
                Visible=False
            ),
            "Abriendo correo en Word",
            progress_callback=progress_callback
        )

        try:
            for shape in doc.InlineShapes:
                if shape.Type in [1, 3, 4, 5]:
                    max_width = 800
                    if shape.Width > max_width:
                        ratio = max_width / shape.Width
                        shape.Width = max_width
                        shape.Height = shape.Height * ratio
        except pywintypes.com_error:
            pass

        ejecutar_com_con_reintentos(
            lambda: doc.ExportAsFixedFormat(OutputFileName=str(pdf_path), ExportFormat=17),
            "Exportando PDF",
            progress_callback=progress_callback
        )

        doc.Close(False)
        doc = None

        eliminar_temporal_con_reintentos(mht_path)

        return pdf_path

    except Exception as error_pdf:
        logger.exception("No se pudo convertir el correo a PDF. Asunto: %s", asunto_limpio)
        raise RuntimeError(f"No se pudo convertir el correo a PDF: {error_pdf}") from error_pdf

    finally:
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass

        eliminar_temporal_con_reintentos(mht_path)


# ============================================================
# Procesamiento backend
# ============================================================

def procesar_exportacion(tipo_exportacion, fecha_inicio, fecha_fin, etiqueta_fecha=None, progress_callback=None):
    """
    Procesa una exportaciÃ³n sin pedir datos por consola.

    tipo_exportacion: "recibidos" o "enviados"
    fecha_inicio: datetime inicial inclusivo
    fecha_fin: datetime final exclusivo
    etiqueta_fecha: texto para nombre de carpeta/Excel. Ej: 2026-05-08 o 2026-05
    progress_callback: funciÃ³n opcional que recibe texto de progreso.
    """

    init(autoreset=True)
    inicializar_personas()
    logger.info(
        "Inicio exportacion. Tipo=%s, fecha_inicio=%s, fecha_fin=%s, etiqueta=%s",
        tipo_exportacion,
        fecha_inicio,
        fecha_fin,
        etiqueta_fecha
    )

    if tipo_exportacion not in ["recibidos", "enviados"]:
        raise ValueError("tipo_exportacion debe ser 'recibidos' o 'enviados'.")

    config = obtener_config_outlook(tipo_exportacion)
    etiqueta_fecha = etiqueta_fecha or fecha_inicio.strftime("%Y-%m-%d")
    mes_name = MESES[fecha_inicio.month]
    anio = str(fecha_inicio.year)

    carpeta_base = (
        Path.home()
        / "Downloads"
        / f"{config['nombre_carpeta']} {mes_name} - {anio}"
        / etiqueta_fecha
    )

    notificar_progreso(progress_callback, "Conectando con Outlook...", 5)

    iniciar_auto_conectar_outlook(progress_callback)

    try:
        outlook_app = ejecutar_com_con_reintentos(
            lambda: win32com.client.gencache.EnsureDispatch("Outlook.Application"),
            "Conectando con Outlook",
            intentos=3,
            espera=2,
            progress_callback=progress_callback
        )
        outlook = ejecutar_com_con_reintentos(
            lambda: outlook_app.GetNamespace("MAPI"),
            "Abriendo perfil de Outlook",
            intentos=3,
            espera=2,
            progress_callback=progress_callback
        )
        carpeta = ejecutar_com_con_reintentos(
            lambda: outlook.GetDefaultFolder(config["folder_id"]),
            "Abriendo carpeta de Outlook",
            intentos=3,
            espera=2,
            progress_callback=progress_callback
        )
    except pywintypes.com_error as error:
        logger.exception("Error conectando con Outlook")
        raise RuntimeError(mensaje_error_com(error)) from error

    items = ejecutar_com_con_reintentos(
        lambda: carpeta.Items,
        "Leyendo correos de Outlook",
        progress_callback=progress_callback
    )
    ejecutar_com_con_reintentos(
        lambda: items.Sort(f"[{config['campo_fecha']}]", config["sort_desc"]),
        "Ordenando correos",
        progress_callback=progress_callback
    )

    items_filtrados = obtener_correos_candidatos(
        items,
        config,
        fecha_inicio,
        fecha_fin,
        progress_callback=progress_callback
    )
    total_filtrados = len(items_filtrados)

    if total_filtrados == 0:
        logger.info("Exportacion sin correos encontrados. Tipo=%s, etiqueta=%s", tipo_exportacion, etiqueta_fecha)
        return ResultadoExportacion(
            exito=False,
            cantidad=0,
            carpeta=carpeta_base,
            excel=None,
            mensaje="No se encontraron correos en el rango especificado."
        )

    notificar_progreso(progress_callback, f"Se encontraron {total_filtrados} correos. Abriendo Word...", 12)
    logger.info("Correos filtrados: %s", total_filtrados)

    word = ejecutar_com_con_reintentos(
        lambda: win32com.client.gencache.EnsureDispatch("Word.Application"),
        "Abriendo Word",
        intentos=3,
        espera=2,
        progress_callback=progress_callback
    )
    word.Visible = False
    try:
        word.DisplayAlerts = 0
    except Exception:
        pass
    try:
        word.ScreenUpdating = False
    except Exception:
        pass
    try:
        word.Options.CheckSpellingAsYouType = False
        word.Options.CheckGrammarAsYouType = False
    except Exception:
        pass

    registros = []
    procesados = 0
    omitidos = 0
    errores = 0

    try:
        for mail in items_filtrados:
            carpeta_correo = None
            try:
                mail_class = ejecutar_com_con_reintentos(
                    lambda: mail.Class,
                    "Leyendo tipo de correo",
                    progress_callback=progress_callback
                )
                if mail_class != 43:
                    omitidos += 1
                    continue

                fecha_py = obtener_fecha_mail(mail, config["atributo_fecha"])
                if not (fecha_inicio <= fecha_py < fecha_fin):
                    omitidos += 1
                    continue

                remitente = ejecutar_com_con_reintentos(
                    lambda: mail.SenderName or "",
                    "Leyendo remitente",
                    progress_callback=progress_callback
                )
                anexos = ejecutar_com_con_reintentos(
                    lambda: mail.Attachments,
                    "Leyendo anexos",
                    progress_callback=progress_callback
                )
                asunto = ejecutar_com_con_reintentos(
                    lambda: mail.Subject or "",
                    "Leyendo asunto",
                    progress_callback=progress_callback
                )
                destinatarios_raw = ejecutar_com_con_reintentos(
                    lambda: mail.To or "",
                    "Leyendo destinatarios",
                    progress_callback=progress_callback
                )
                cc = ejecutar_com_con_reintentos(
                    lambda: mail.CC or "",
                    "Leyendo copias",
                    progress_callback=progress_callback
                )

                if tipo_exportacion == "recibidos" and remitente in REMITENTES_OMITIDOS:
                    omitidos += 1
                    continue

                procesados += 1
                porcentaje = 15 + int((procesados / total_filtrados) * 75)
                notificar_progreso(
                    progress_callback,
                    f"Procesando correo {procesados} de {total_filtrados}: {asunto[:80]}",
                    porcentaje
                )

                nombre_carpeta_base = obtener_nombre_para_carpeta(
                    tipo_exportacion,
                    remitente,
                    destinatarios_raw
                )

                id_correo = fecha_py.strftime("%H%M%S") + "_" + limpiar_acortar_remitentes(nombre_carpeta_base)
                carpeta_nombre = id_correo[:100]

                carpeta_base.mkdir(parents=True, exist_ok=True)
                carpeta_correo = carpeta_base / carpeta_nombre

                contador = 1
                carpeta_original = carpeta_correo
                while carpeta_correo.exists():
                    contador += 1
                    carpeta_correo = Path(str(carpeta_original)[:95] + f"_{contador}")

                os.makedirs(carpeta_correo, exist_ok=True)

                asunto_limpio = limpiar_texto(asunto) or "Sin asunto"
                convertir_correo_a_pdf(
                    mail,
                    word,
                    carpeta_correo,
                    asunto_limpio,
                    progress_callback=progress_callback
                )

                remitente_filtrado = nombres_conocidos_rem(remitente, fecha_py)
                cargo, dependencia = obtener_info_remitente(remitente, fecha_py)

                destinatarios_cortos = cut_nombres_destinatarios(destinatarios_raw)
                destinatarios_final, cargos = obtener_info_destinatarios(destinatarios_cortos, fecha_py)

                cc_filtrado = nombres_conocidos_cc(str(cc), fecha_py)

                lista_anexos = obtener_anexos(anexos, carpeta_correo)
                cant_anexos = len(lista_anexos)

                observaciones = (
                    "No contiene anexos"
                    if cant_anexos == 0
                    else f"Anexa {cant_anexos} documento(s)"
                )

                registros.append({
                    "Tipo del Documento": "CORREO INSTITUCIONAL",
                    "Fecha del Documento": fecha_py.strftime("%Y-%m-%d"),
                    "Remitente": nompropio_python(remitente_filtrado),
                    "Cargo": cargo,
                    "Facultad/Dependencia": dependencia,
                    "Destinatario": nompropio_python(destinatarios_final),
                    "Empresa/Cargo": cargos,
                    "Asunto": nompropio_python(asunto),
                    "Con Copia": cc_filtrado,
                    "Observaciones": observaciones
                })

            except Exception as e:
                errores += 1
                logger.exception("Error procesando un correo. Procesados=%s", procesados)
                if carpeta_correo:
                    limpiar_carpeta_si_vacia(carpeta_correo)
                notificar_progreso(progress_callback, f"Error procesando un correo: {e}")

    finally:
        try:
            word.Quit()
        except Exception:
            pass

    if not registros:
        logger.info("Exportacion sin registros validos. Tipo=%s, etiqueta=%s", tipo_exportacion, etiqueta_fecha)
        return ResultadoExportacion(
            exito=False,
            cantidad=0,
            carpeta=carpeta_base,
            excel=None,
            mensaje=(
                "No se encontraron correos validos para exportar. "
                f"Candidatos: {total_filtrados}. Omitidos: {omitidos}. Errores: {errores}."
            ),
            total=total_filtrados,
            errores=errores,
            omitidos=omitidos
        )

    notificar_progreso(progress_callback, "Creando archivo Excel...", 94)

    ruta_excel = exportar_excel(registros, carpeta_base, etiqueta_fecha, tipo_exportacion)
    logger.info(
        "Exportacion terminada. Registros=%s, omitidos=%s, errores=%s, carpeta=%s, excel=%s",
        len(registros),
        omitidos,
        errores,
        carpeta_base,
        ruta_excel
    )

    notificar_progreso(progress_callback, "Exportacion completada.", 100)

    return ResultadoExportacion(
        exito=True,
        cantidad=len(registros),
        carpeta=carpeta_base,
        excel=ruta_excel,
        mensaje=(
            f"Exportacion completada. Exportados: {len(registros)}. "
            f"Omitidos: {omitidos}. Errores: {errores}."
        ),
        total=total_filtrados,
        errores=errores,
        omitidos=omitidos
    )


# ============================================================
# Modo consola bÃ¡sico, solo para compatibilidad/pruebas
# ============================================================

def pedir_fecha():
    while True:
        try:
            dia = input("DÃ­a (1-31): ")
            mes = input("Mes (1-12): ")
            anio = input("AÃ±o (4 dÃ­gitos): ")
            return obtener_rango_dia(dia, mes, anio)
        except ValueError:
            print(Fore.RED + "\nFecha invÃ¡lida. Intente nuevamente.\n")


def pedir_tipo_exportacion():
    while True:
        print("Â¿QuÃ© desea exportar?")
        print("1. Correos recibidos")
        print("2. Correos enviados")

        opcion = input("\nSeleccione una opciÃ³n (1 o 2): ").strip()

        if opcion == "1":
            return "recibidos"
        if opcion == "2":
            return "enviados"

        print(Fore.RED + "\nOpciÃ³n invÃ¡lida. Escriba 1 o 2.\n")


def procesar():
    print("=== Exportador de Correos y Anexos ===", end="\n\n")
    tipo_exportacion = pedir_tipo_exportacion()
    f_inicio, f_fin, etiqueta, _mes_name, _anio = pedir_fecha()

    resultado = procesar_exportacion(
        tipo_exportacion=tipo_exportacion,
        fecha_inicio=f_inicio,
        fecha_fin=f_fin,
        etiqueta_fecha=etiqueta,
        progress_callback=print
    )

    print("\n" + resultado.mensaje)
    if resultado.exito and resultado.carpeta:
        input("\nPresione ENTER para abrir la carpeta.")
        os.startfile(resultado.carpeta)
    else:
        input("\nPresione ENTER para continuar.")


if __name__ == "__main__":
    while True:
        procesar()
